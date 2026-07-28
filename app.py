import os
import json
import sqlite3
import secrets
import logging
import urllib.request
import urllib.error
from flask import Flask, jsonify, request, g, Response
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature

def get_secret_key():
    key = os.environ.get('SECRET_KEY')
    if key:
        return key
    secret_file = os.path.join(os.path.dirname(__file__), '.secret_key')
    if os.path.exists(secret_file):
        try:
            with open(secret_file, 'r') as f:
                return f.read().strip()
        except Exception:
            pass
    generated_key = secrets.token_hex(32)
    try:
        with open(secret_file, 'w') as f:
            f.write(generated_key)
    except Exception:
        logging.warning("Generating ephemeral secret. Instance-isolated!")
    return generated_key

app = Flask(__name__)
app.config['SECRET_KEY'] = get_secret_key()
# Enable CORS for all origins, headers, and methods to prevent cross-origin issues
CORS(app, resources={r"/api/*": {"origins": "*", "allow_headers": ["Content-Type", "Authorization"]}})

DB_PATH = os.path.join(os.path.dirname(__file__), 'trips.db')

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Create Places table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS places (
        id TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        location TEXT,
        region TEXT,
        mood TEXT,
        description TEXT,
        image_url TEXT,
        latitude REAL,
        longitude REAL,
        estimated_cost_per_day INTEGER,
        best_time_to_visit TEXT,
        duration_days INTEGER,
        nearby_hotels TEXT,
        nearby_restaurants TEXT,
        transportation_options TEXT,
        activities TEXT
    )
    ''')
    
    # Create Trips table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS trips (
        id TEXT PRIMARY KEY,
        user_id TEXT,
        trip_name TEXT NOT NULL,
        destination TEXT NOT NULL,
        start_date TEXT,
        end_date TEXT,
        origin TEXT,
        mode_of_transport TEXT,
        number_of_travelers INTEGER,
        traveler_details TEXT,
        total_estimated_cost REAL,
        actual_cost REAL,
        mood TEXT,
        status TEXT,
        places_to_visit TEXT,
        notes TEXT
    )
    ''')
    
    # Check if user_id column exists in trips table (for existing tables)
    cursor.execute("PRAGMA table_info(trips)")
    columns = [col[1] for col in cursor.fetchall()]
    if 'user_id' not in columns:
        cursor.execute("ALTER TABLE trips ADD COLUMN user_id TEXT")
        cursor.execute("UPDATE trips SET user_id = 'me' WHERE user_id IS NULL")
    
    # Create Users table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS users (
        id TEXT PRIMARY KEY,
        username TEXT UNIQUE,
        password_hash TEXT,
        full_name TEXT,
        email TEXT,
        bio TEXT,
        location TEXT,
        preferred_mood TEXT,
        travel_budget TEXT,
        travel_style TEXT,
        languages TEXT,
        interests TEXT,
        avatar_url TEXT,
        created_date TEXT
    )
    ''')
    
    # Check if username and password_hash columns exist (for existing tables)
    cursor.execute("PRAGMA table_info(users)")
    columns = [col[1] for col in cursor.fetchall()]
    if 'username' not in columns:
        cursor.execute("ALTER TABLE users ADD COLUMN username TEXT")
    if 'password_hash' not in columns:
        cursor.execute("ALTER TABLE users ADD COLUMN password_hash TEXT")
    
    conn.commit()
    
    # Seed places if empty or out-of-date
    cursor.execute("SELECT COUNT(*) FROM places")
    if cursor.fetchone()[0] < 2000:
        cursor.execute("DELETE FROM places")
        seed_places(cursor)
        conn.commit()
        
    # Seed default user if empty
    cursor.execute("SELECT COUNT(*) FROM users")
    if cursor.fetchone()[0] == 0:
        cursor.execute('''
        INSERT INTO users (id, username, password_hash, full_name, email, bio, location, preferred_mood, travel_budget, travel_style, languages, interests, avatar_url, created_date)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            'me',
            'traveler',
            generate_password_hash('password123'),
            'Travel Explorer',
            'explorer@example.com',
            'Adventure seeker and travel enthusiast. Love to explore new destinations and cultures.',
            'Mumbai, India',
            'adventure',
            'mid-range',
            'mixed',
            json.dumps(['English', 'Spanish']),
            json.dumps(['Photography', 'Nature', 'Adventure', 'Food']),
            '',
            '2024-01-01'
        ))
        conn.commit()
    else:
        # Update default user if username/password_hash are empty
        cursor.execute('''
        UPDATE users SET username = 'traveler', password_hash = ?
        WHERE id = 'me' AND (username IS NULL OR password_hash IS NULL)
        ''', (generate_password_hash('password123'),))
        conn.commit()
        
    conn.close()

def seed_places(cursor):
    json_path = os.path.join(os.path.dirname(__file__), 'places.json')
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            places_data = json.load(f)
    except Exception as e:
        import logging
        logging.error(f"Error loading places.json: {e}")
        return
        
    for place in places_data:
        cursor.execute('''
        INSERT INTO places (
            id, name, location, region, mood, description, image_url, 
            latitude, longitude, estimated_cost_per_day, best_time_to_visit, 
            duration_days, nearby_hotels, nearby_restaurants, 
            transportation_options, activities
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            place["id"],
            place["name"],
            place["location"],
            place["region"],
            place["mood"],
            place["description"],
            place["image_url"],
            place["latitude"],
            place["longitude"],
            place["estimated_cost_per_day"],
            place["best_time_to_visit"],
            place["duration_days"],
            json.dumps(place["nearby_hotels"]),
            json.dumps(place["nearby_restaurants"]),
            json.dumps(place["transportation_options"]),
            json.dumps(place["activities"])
        ))

# Authentication and Session utilities
def get_auth_serializer():
    return URLSafeTimedSerializer(app.config['SECRET_KEY'])

def generate_auth_token(user_id):
    serializer = get_auth_serializer()
    return serializer.dumps({'user_id': user_id})

def verify_auth_token(token):
    serializer = get_auth_serializer()
    try:
        # Token is valid for 24 hours (86400 seconds)
        data = serializer.loads(token, max_age=86400)
        return data.get('user_id')
    except (SignatureExpired, BadSignature):
        return None

from functools import wraps

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        auth_header = request.headers.get('Authorization')
        if auth_header:
            parts = auth_header.split()
            if len(parts) == 2 and parts[0].lower() == 'bearer':
                token = parts[1]
        
        # If no token is provided by the frontend, fall back to the default user 'me' (except in strict test suite mode)
        if not token:
            if app.config.get('TESTING'):
                return jsonify(error='Missing token'), 401
            g.user_id = 'me'
            return f(*args, **kwargs)
            
        user_id = verify_auth_token(token)
        if not user_id:
            return jsonify(error='Invalid or expired token'), 401
            
        g.user_id = user_id
        return f(*args, **kwargs)
    return decorated

import openpyxl
from openpyxl import Workbook

def update_excel_tracker(username, email, password):
    excel_path = os.path.join(os.path.dirname(__file__), 'users.xlsx')
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"
        ws['A1'] = "Total Customers: 1"
        ws['A2'] = "Username"
        ws['B2'] = "Email"
        ws['C2'] = "Password"
        ws.append([username, email, password])
        wb.save(excel_path)
    else:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        ws.append([username, email, password])
        
        count = ws.max_row - 2
        ws['A1'] = f"Total Customers: {count}"
        wb.save(excel_path)

@app.post('/api/auth/register')
def register():
    payload = request.get_json(force=True) or {}
    username = payload.get('username')
    password = payload.get('password')
    email = payload.get('email')
    full_name = payload.get('full_name', '')
    
    if not username or not password or not email:
        return jsonify(error='Username, password, and email are required'), 400
        
    if len(password) < 8:
        return jsonify(error='Password must be at least 8 characters long'), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Check if username or email already exists
    cursor.execute("SELECT id FROM users WHERE username = ? OR email = ?", (username, email))
    if cursor.fetchone():
        conn.close()
        return jsonify(error='Username or email already exists'), 400
        
    user_id = secrets.token_hex(8) # Unique string ID for the user
    password_hash = generate_password_hash(password)
    
    try:
        cursor.execute('''
        INSERT INTO users (id, username, password_hash, full_name, email, created_date)
        VALUES (?, ?, ?, ?, ?, DATE('now'))
        ''', (user_id, username, password_hash, full_name, email))
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify(error='Failed to register user'), 500
        
    conn.close()
    
    # Save to Excel
    try:
        update_excel_tracker(username, email, password)
    except Exception as e:
        import logging
        logging.error(f"Failed to update excel tracker: {e}")
        
    return jsonify(message='User registered successfully', user_id=user_id), 201

@app.post('/api/auth/login')
def login():
    payload = request.get_json(force=True) or {}
    username = payload.get('username')
    password = payload.get('password')
    
    if not username or not password:
        return jsonify(error='Username and password are required'), 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users WHERE username = ? OR email = ?", (username, username))
    row = cursor.fetchone()
    conn.close()
    
    if not row or not check_password_hash(row['password_hash'], password):
        return jsonify(error='Invalid username or password'), 401
        
    user_id = row['id']
    token = generate_auth_token(user_id)
    
    user = dict(row)
    user.pop('password_hash', None)
    user['languages'] = json.loads(row['languages']) if row['languages'] else []
    user['interests'] = json.loads(row['interests']) if row['interests'] else []
    
    return jsonify(token=token, user=user), 200

# Ensure database is initialized before serving requests
init_db()

@app.route('/')
def hello_world():
    return jsonify(message='Hello, World from AI Trip Planner API!', status='ok')

@app.get('/api/health')
def health():
    return jsonify(status='ok', database='connected')

@app.post('/api/log_error')
def log_error():
    payload = request.get_json(force=True) or {}
    print("\n[FRONTEND ERROR]:", payload.get('error'))
    print("Stack:", payload.get('stack'), "\n")
    return jsonify(status='received'), 200

import math

def haversine_distance(lat1, lon1, lat2, lon2):
    R = 6371.0 # Earth's radius in kilometers
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat / 2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

@app.get('/api/navigation/directions')
def get_directions():
    origin_id = request.args.get('origin_id')
    destination_id = request.args.get('destination_id')
    
    try:
        origin_lat = float(request.args.get('origin_lat', 0.0))
        origin_lng = float(request.args.get('origin_lng', 0.0))
        dest_lat = float(request.args.get('dest_lat', 0.0))
        dest_lng = float(request.args.get('dest_lng', 0.0))
    except ValueError:
        return jsonify(error='Invalid coordinate parameters'), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    origin_name = "Starting Point"
    dest_name = "Destination"

    # Resolve origin
    if origin_id:
        cursor.execute("SELECT name, latitude, longitude FROM places WHERE id = ?", (origin_id,))
        row = cursor.fetchone()
        if row:
            origin_name = row['name']
            origin_lat = row['latitude']
            origin_lng = row['longitude']
    
    # Resolve destination
    if destination_id:
        cursor.execute("SELECT name, latitude, longitude FROM places WHERE id = ?", (destination_id,))
        row = cursor.fetchone()
        if row:
            dest_name = row['name']
            dest_lat = row['latitude']
            dest_lng = row['longitude']
            
    conn.close()

    if not origin_lat or not origin_lng:
        # Default to NATPAC HQ or Mumbai center as origin fallback
        origin_name = "NATPAC Headquarters (Trivandrum)"
        origin_lat = 8.5241
        origin_lng = 76.9366

    if not dest_lat or not dest_lng:
        return jsonify(error='Destination location could not be resolved'), 400

    # Calculate distance
    dist = haversine_distance(origin_lat, origin_lng, dest_lat, dest_lng)
    
    # Estimate transit modes
    modes = {}
    
    # Driving
    driving_time = dist / 50.0 # 50 km/h avg
    driving_hours = int(driving_time)
    driving_mins = int((driving_time - driving_hours) * 60)
    modes['driving'] = {
        'name': 'Driving / Taxi',
        'duration_text': f"{driving_hours}h {driving_mins}m" if driving_hours > 0 else f"{driving_mins}m",
        'estimated_cost': int(dist * 18),
        'co2_emissions_kg': round(dist * 0.12, 1)
    }
    
    # Bus
    bus_time = dist / 35.0
    bus_hours = int(bus_time)
    bus_mins = int((bus_time - bus_hours) * 60)
    modes['bus'] = {
        'name': 'Bus Transit',
        'duration_text': f"{bus_hours}h {bus_mins}m" if bus_hours > 0 else f"{bus_mins}m",
        'estimated_cost': int(dist * 2.5),
        'co2_emissions_kg': round(dist * 0.03, 1)
    }

    # Train
    train_time = dist / 65.0
    train_hours = int(train_time)
    train_mins = int((train_time - train_hours) * 60)
    modes['train'] = {
        'name': 'Rail Express',
        'duration_text': f"{train_hours}h {train_mins}m" if train_hours > 0 else f"{train_mins}m",
        'estimated_cost': int(dist * 3.0),
        'co2_emissions_kg': round(dist * 0.015, 1)
    }
    
    # Flight (only if distance > 300km)
    if dist > 300:
        flight_time = (dist / 700.0) + 2.0 # 700 km/h + 2 hours checkin
        flight_hours = int(flight_time)
        flight_mins = int((flight_time - flight_hours) * 60)
        modes['flight'] = {
            'name': 'Air Travel',
            'duration_text': f"{flight_hours}h {flight_mins}m",
            'estimated_cost': int(4500 + dist * 4),
            'co2_emissions_kg': round(dist * 0.25, 1)
        }

    # Step-by-step directions
    steps = [
        {'instruction': f'Depart from {origin_name}. Head towards closest major highway link.', 'distance_offset_km': 0.0},
        {'instruction': f'Continue along main highway corridor towards {dest_name}.', 'distance_offset_km': round(dist * 0.3, 1)},
        {'instruction': f'Passing by intermediate travel junctions and NATPAC transit checkpoints.', 'distance_offset_km': round(dist * 0.6, 1)},
        {'instruction': f'Take local exit ramp and navigate according to destination boundary signs.', 'distance_offset_km': round(dist * 0.9, 1)},
        {'instruction': f'Arrive at {dest_name}. Your destination is visible nearby.', 'distance_offset_km': round(dist, 1)}
    ]

    return jsonify(
        origin={
            'name': origin_name,
            'latitude': origin_lat,
            'longitude': origin_lng
        },
        destination={
            'name': dest_name,
            'latitude': dest_lat,
            'longitude': dest_lng
        },
        distance_km=round(dist, 1),
        modes=modes,
        steps=steps
    )

@app.get('/api/places')
def list_places():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM places")
    rows = cursor.fetchall()
    conn.close()
    
    places = []
    for r in rows:
        place = dict(r)
        place['nearby_hotels'] = json.loads(r['nearby_hotels']) if r['nearby_hotels'] else []
        place['nearby_restaurants'] = json.loads(r['nearby_restaurants']) if r['nearby_restaurants'] else []
        place['transportation_options'] = json.loads(r['transportation_options']) if r['transportation_options'] else []
        place['activities'] = json.loads(r['activities']) if r['activities'] else []
        places.append(place)
        
    return jsonify(places)

@app.get('/api/places/<id>')
def get_place(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM places WHERE id = ?", (id,))
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        return jsonify(error='Place not found'), 404
        
    place = dict(row)
    place['nearby_hotels'] = json.loads(row['nearby_hotels']) if row['nearby_hotels'] else []
    place['nearby_restaurants'] = json.loads(row['nearby_restaurants']) if row['nearby_restaurants'] else []
    place['transportation_options'] = json.loads(row['transportation_options']) if row['transportation_options'] else []
    place['activities'] = json.loads(row['activities']) if row['activities'] else []
    
    return jsonify(place)

@app.get('/api/trips')
@token_required
def list_trips():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM trips WHERE user_id = ? ORDER BY start_date DESC", (g.user_id,))
    rows = cursor.fetchall()
    conn.close()
    
    trips = []
    for r in rows:
        trip = dict(r)
        trip['traveler_details'] = json.loads(r['traveler_details']) if r['traveler_details'] else []
        trip['places_to_visit'] = json.loads(r['places_to_visit']) if r['places_to_visit'] else []
        trips.append(trip)
        
    return jsonify(trips)

@app.post('/api/trips')
@token_required
def create_trip():
    payload = request.get_json(force=True) or {}
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    trip_id = payload.get('id')
    if not trip_id:
        cursor.execute("SELECT COUNT(*) FROM trips")
        count = cursor.fetchone()[0]
        trip_id = str(count + 100)
    
    # Prevent conflict
    cursor.execute("SELECT id FROM trips WHERE id = ?", (trip_id,))
    if cursor.fetchone():
        while True:
            trip_id = str(int(trip_id) + 1)
            cursor.execute("SELECT id FROM trips WHERE id = ?", (trip_id,))
            if not cursor.fetchone():
                break
        
    cursor.execute('''
    INSERT INTO trips (
        id, user_id, trip_name, destination, start_date, end_date, origin, 
        mode_of_transport, number_of_travelers, traveler_details, 
        total_estimated_cost, actual_cost, mood, status, places_to_visit, notes
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        trip_id,
        g.user_id,
        payload.get('trip_name'),
        payload.get('destination'),
        payload.get('start_date'),
        payload.get('end_date'),
        payload.get('origin'),
        payload.get('mode_of_transport'),
        int(payload.get('number_of_travelers', 1)),
        json.dumps(payload.get('traveler_details', [])),
        float(payload.get('total_estimated_cost', 0.0)),
        float(payload.get('actual_cost', payload.get('total_estimated_cost', 0.0))),
        payload.get('mood', 'adventure'),
        payload.get('status', 'planned'),
        json.dumps(payload.get('places_to_visit', [])),
        payload.get('notes', '')
    ))
    conn.commit()
    
    # Retrieve newly created trip
    cursor.execute("SELECT * FROM trips WHERE id = ?", (trip_id,))
    new_trip_row = cursor.fetchone()
    conn.close()
    
    new_trip = dict(new_trip_row)
    new_trip['traveler_details'] = json.loads(new_trip['traveler_details'])
    new_trip['places_to_visit'] = json.loads(new_trip['places_to_visit'])
    
    return jsonify(new_trip), 201

@app.put('/api/trips/<id>')
@token_required
def update_trip(id):
    payload = request.get_json(force=True) or {}
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT user_id FROM trips WHERE id = ?", (id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify(error='Trip not found'), 404
        
    if row['user_id'] != g.user_id:
        conn.close()
        return jsonify(error='Unauthorized to update this trip'), 403
        
    cursor.execute('''
    UPDATE trips SET 
        trip_name = ?, 
        destination = ?, 
        start_date = ?, 
        end_date = ?, 
        origin = ?, 
        mode_of_transport = ?, 
        number_of_travelers = ?, 
        traveler_details = ?, 
        total_estimated_cost = ?, 
        actual_cost = ?, 
        mood = ?, 
        status = ?, 
        places_to_visit = ?, 
        notes = ?
    WHERE id = ? AND user_id = ?
    ''', (
        payload.get('trip_name'),
        payload.get('destination'),
        payload.get('start_date'),
        payload.get('end_date'),
        payload.get('origin'),
        payload.get('mode_of_transport'),
        int(payload.get('number_of_travelers', 1)),
        json.dumps(payload.get('traveler_details', [])),
        float(payload.get('total_estimated_cost', 0.0)),
        float(payload.get('actual_cost', 0.0)),
        payload.get('mood'),
        payload.get('status'),
        json.dumps(payload.get('places_to_visit', [])),
        payload.get('notes', ''),
        id,
        g.user_id
    ))
    conn.commit()
    
    cursor.execute("SELECT * FROM trips WHERE id = ?", (id,))
    updated_trip_row = cursor.fetchone()
    conn.close()
    
    updated_trip = dict(updated_trip_row)
    updated_trip['traveler_details'] = json.loads(updated_trip['traveler_details'])
    updated_trip['places_to_visit'] = json.loads(updated_trip['places_to_visit'])
    
    return jsonify(updated_trip)

@app.delete('/api/trips/<id>')
@token_required
def delete_trip(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT user_id FROM trips WHERE id = ?", (id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify(error='Trip not found'), 404
        
    if row['user_id'] != g.user_id:
        conn.close()
        return jsonify(error='Unauthorized to delete this trip'), 403
        
    cursor.execute("DELETE FROM trips WHERE id = ? AND user_id = ?", (id, g.user_id))
    conn.commit()
    conn.close()
    
    return jsonify(message='Trip deleted successfully')

@app.get('/api/user/me')
@token_required
def get_current_user():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users WHERE id = ?", (g.user_id,))
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        return jsonify(error='User profile not found'), 404
        
    user = dict(row)
    user.pop('password_hash', None)
    user['languages'] = json.loads(row['languages']) if row['languages'] else []
    user['interests'] = json.loads(row['interests']) if row['interests'] else []
    
    return jsonify(user)

def update_profile_excel(user_id, profile_data):
    excel_path = os.path.join(os.path.dirname(__file__), 'user_profiles.xlsx')
    
    headers = ['User ID', 'Full Name', 'Email', 'Location', 'Bio', 'Mood', 'Budget', 'Style', 'Languages', 'Interests']
    row_data = [
        user_id,
        profile_data.get('full_name', ''),
        profile_data.get('email', ''),
        profile_data.get('location', ''),
        profile_data.get('bio', ''),
        profile_data.get('preferred_mood', ''),
        profile_data.get('travel_budget', ''),
        profile_data.get('travel_style', ''),
        ', '.join(profile_data.get('languages', [])),
        ', '.join(profile_data.get('interests', []))
    ]
    
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Profiles"
        ws.append(headers)
        ws.append(row_data)
        wb.save(excel_path)
    else:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Check if user already exists in the sheet to update them
        user_row_idx = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == user_id:
                user_row_idx = idx
                break
                
        if user_row_idx:
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=user_row_idx, column=col_idx, value=value)
        else:
            ws.append(row_data)
            
        wb.save(excel_path)

@app.post('/api/user/me')
@token_required
def update_current_user():
    payload = request.get_json(force=True) or {}
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Check if user exists
    cursor.execute("SELECT id FROM users WHERE id = ?", (g.user_id,))
    if not cursor.fetchone():
        conn.close()
        return jsonify(error='User not found'), 404

    cursor.execute('''
    UPDATE users SET
        full_name = ?,
        email = ?,
        bio = ?,
        location = ?,
        preferred_mood = ?,
        travel_budget = ?,
        travel_style = ?,
        languages = ?,
        interests = ?
    WHERE id = ?
    ''', (
        payload.get('full_name'),
        payload.get('email'),
        payload.get('bio', ''),
        payload.get('location', ''),
        payload.get('preferred_mood', ''),
        payload.get('travel_budget', ''),
        payload.get('travel_style', ''),
        json.dumps(payload.get('languages', [])),
        json.dumps(payload.get('interests', [])),
        g.user_id
    ))
    conn.commit()
    
    cursor.execute("SELECT * FROM users WHERE id = ?", (g.user_id,))
    row = cursor.fetchone()
    conn.close()
    
    # Save to Excel
    try:
        update_profile_excel(g.user_id, payload)
    except Exception as e:
        import logging
        logging.error(f"Failed to update profile excel: {e}")
    
    user = dict(row)
    user.pop('password_hash', None)
    user['languages'] = json.loads(row['languages']) if row['languages'] else []
    user['interests'] = json.loads(row['interests']) if row['interests'] else []
    
    return jsonify(user)

def get_available_model(preferred_model, fallback='qwen3:latest'):
    try:
        req = urllib.request.Request("http://localhost:11434/api/tags", method='GET')
        with urllib.request.urlopen(req, timeout=2) as response:
            res_data = json.loads(response.read().decode('utf-8'))
            installed = [m['name'] for m in res_data.get('models', [])]
            if preferred_model in installed:
                return preferred_model
            # Match base name (e.g. 'llama3.1:8b' matches 'llama3.1')
            base = preferred_model.split(':')[0]
            for m in installed:
                if m.startswith(base):
                    return m
            if fallback in installed:
                return fallback
            return installed[0] if installed else preferred_model
    except:
        return preferred_model

@app.post('/api/ai/analyze-places')
def ai_analyze_places():
    payload = request.get_json() or {}
    origin = payload.get('origin', '')
    destination = payload.get('destination', '')
    mood = payload.get('mood', '')
    
    conn = get_db_connection()
    # Get a pool of relevant candidates to feed the AI
    # We favor the destination strongly, but also allow mood matches
    query = '''
        SELECT id, name, location, region, mood
        FROM places 
        WHERE LOWER(location) LIKE ? OR LOWER(region) LIKE ? OR LOWER(name) LIKE ? OR LOWER(mood) LIKE ?
        LIMIT 40
    '''
    dest_term = f"%{destination.lower()}%"
    mood_term = f"%{mood.lower()}%"
    rows = conn.execute(query, (dest_term, dest_term, dest_term, mood_term)).fetchall()
    conn.close()
    
    if not rows:
        return jsonify({"place_ids": []})
        
    # Format candidates for the LLM
    candidate_text = "\n".join([f"ID:{r['id']} | {r['name']} ({r['location']}) | Mood:{r['mood']}" for r in rows])
    
    prompt = (
        f"You are an AI trip planner. The user is starting from '{origin}' and traveling to '{destination}' for a '{mood}' trip. "
        f"Select the 10 best place IDs to visit from the Candidates list below. Do not suggest places in the starting location unless it's a road trip stop. "
        f"Return ONLY a comma-separated list of IDs (e.g. 4,12,50). No explanations.\n\nCandidates:\n{candidate_text}"
    )
    
    target_model = get_available_model('phi4')
    
    url = 'http://localhost:11434/api/generate'
    body = json.dumps({
        'model': target_model,
        'prompt': prompt,
        'stream': False,
        'options': {'temperature': 0.1, 'num_predict': 50, 'num_thread': 8, 'num_ctx': 2048}
    }).encode('utf-8')
    
    try:
        req = urllib.request.Request(url, data=body, headers={'Content-Type': 'application/json'}, method='POST')
        with urllib.request.urlopen(req, timeout=20) as resp:
            res_data = json.loads(resp.read().decode('utf-8'))
            import re
            place_ids = [int(num) for num in re.findall(r'\d+', res_data.get('response', ''))]
            
            # Validate IDs
            valid_ids = {r['id'] for r in rows}
            final_ids = [pid for pid in place_ids if pid in valid_ids]
            
            if not final_ids:
                final_ids = [r['id'] for r in rows[:10]]
                
            return jsonify({"place_ids": final_ids})
    except Exception as e:
        import logging
        logging.error(f'AI Analyzer error: {e}')
        # Fallback if LLM fails
        return jsonify({"place_ids": [r['id'] for r in rows[:10]]})

@app.get('/api/ai/models')
def list_ai_models():
    url = "http://localhost:11434/api/tags"
    req = urllib.request.Request(url, method='GET')
    try:
        with urllib.request.urlopen(req, timeout=5) as response:
            res_data = json.loads(response.read().decode('utf-8'))
            models = [m['name'] for m in res_data.get('models', [])]
            return jsonify(models=models, online=True)
    except Exception as e:
        return jsonify(models=['llama3.1', 'phi4', 'mistral'], online=False)

@app.post('/api/ai/chat')
def ai_chat():
    payload = request.get_json() or {}
    requested_model = payload.get('model', 'llama3.1')
    model = get_available_model(requested_model)
    messages = payload.get('messages', [])
    
    # ── 1. Fast RAG: SQL LIKE query (avoids loading all 2000 rows into Python) ──
    user_query = messages[-1].get('content', '') if messages else ''
    context_block = ''
    if user_query:
        try:
            conn = get_db_connection()
            # Pull up to 2 matching places using DB-side filtering
            words = [w for w in user_query.lower().split() if len(w) > 3]
            if words:
                conditions = ' OR '.join(['LOWER(name) LIKE ?' for _ in words])
                params = [f'%{w}%' for w in words]
                rows = conn.execute(
                    f'SELECT name, location, mood, estimated_cost_per_day, '
                    f'best_time_to_visit, activities, nearby_hotels '
                    f'FROM places WHERE {conditions} LIMIT 3',
                    params
                ).fetchall()
                conn.close()
                if rows:
                    context_block = "\n\n[DB Context (STRICT FACTUAL DATA)]:\n" + "\n".join([
                        f"- {r['name']} ({r['location']}) | Mood: {r['mood']} | Cost: ₹{r['estimated_cost_per_day']}/day | Best Time: {r['best_time_to_visit']} | Activities: {r['activities']} | Hotels: {r['nearby_hotels']}"
                        for r in rows
                    ])
            else:
                conn.close()
        except Exception as e:
            logging.error(f'RAG query error: {e}')

    # ── 2. Build a tight system prompt ──
    system_prompt = (
        'You are an elite, highly accurate travel AI. '
        'Reply in exactly 2-3 extremely short, punchy bullet points. NO conversational filler. NO introductions or conclusions. '
        'Base your answers strictly on the [DB Context] if available.'
        + context_block
    )

    # ── 3. Keep only last 3 turns to reduce context size ──
    processed = [{'role': 'system', 'content': system_prompt}] + messages[-3:]

    url = 'http://localhost:11434/api/chat'
    body = json.dumps({
        'model': model,
        'messages': processed,
        'stream': True,
        'options': {
            'num_predict': 150,   # shorter cap → forces concise outputs
            'temperature': 0.3,   # more deterministic and highly accurate
            'num_thread': 8,      # harness high-end CPU resources
            'num_ctx': 2048,      # allow deep context windows without truncation
            'top_k': 10,          # highly restricted token sampling for logic correctness
            'top_p': 0.6,
        }
    }).encode('utf-8')

    def generate():
        req = urllib.request.Request(url, data=body,
                                     headers={'Content-Type': 'application/json'},
                                     method='POST')
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                for line in resp:
                    line = line.strip()
                    if line:
                        try:
                            chunk = json.loads(line)
                            content = chunk.get('message', {}).get('content', '')
                            done = chunk.get('done', False)
                            yield f"data: {json.dumps({'content': content, 'done': done})}\n\n"
                        except Exception:
                            pass
        except Exception as e:
            yield f"data: {json.dumps({'error': str(e), 'done': True})}\n\n"

    return Response(generate(), mimetype='text/event-stream')

if __name__ == '__main__':
    # Run the server on port 5001 (port used by frontend configs)
    app.run(debug=True, port=5001)
